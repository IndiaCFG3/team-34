
from openpyxl import Workbook
from openpyxl import load_workbook
import random

class dataAsset():
    scoreFromExcelSheet = 0.00
    totalPopulation = 0
    success = 0
    def __init__(self):
        scoreFromExcelSheet = 0.00
        totalPopulation = 0
        success = 0


    def readExcel(self, fileN):
        worksheet = load_workbook(filename=fileN)
        return worksheet.active
        # print(sheet['A'])
        # for val in sheet['A']:
        #     print(val.value)

    # def makeExcel(self, fileN):
    #     workbook = Workbook()
    #     sheet = workbook.active
    #     for i in range(1,10):
    #         k = random.randint(0,1)
    #         string = "C"+str(i)
    #         # print(string)
    #         if k==1:
    #             sheet[string] = "Y"
    #         else:
    #             sheet[string] = "N"                
    #         print(sheet[string].value,string)

    #     sheet["A1"] = "hello"
    #     sheet["A2"] = "dfsd"
    #     sheet["B1"] = "world!"

    #     workbook.save(filename=fileN)

        
    def getSuccessStats(self, fileN):
        sheet = self.readExcel(fileN)
        for success in sheet['C']:
            self.totalPopulation +=1
            if(success.value == "Y" or success.value == "y"):
                self.success+=1
        self.scoreFromExcelSheet = float(self.success/self.totalPopulation)*100
        print(self.scoreFromExcelSheet)
        # print('A')


dA = dataAsset()
dA.makeExcel("hi.xlsx")
# dA.readExcel("hi.xlsx")
dA.getSuccessStats("hi.xlsx")